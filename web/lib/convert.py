"""
Convert supplementary data into other formats, e.g., Excel to TSV
"""
import os
import logging
from .config import use_config

log = logging.getLogger(__name__)

if os.getenv('DEBUG') or os.getenv('DEBUG_DATA'):
    log.setLevel(logging.DEBUG)

DEFAULT_DELIM = '\t'


def trim_data_area(sheet):
    """
    Trim "data area" for given sheet, to exclude non-data cells

    In other words, remove columns and rows that are only considered non-empty
    because they contain formatting, starting from right-most column
    """
    for col in range(sheet.max_column, 1, -1):
        values = [
            sheet.cell(row, col).value is not None
            for row in range(1, sheet.max_row + 1)
        ]
        if sum(values):
            log.debug("Last column of data area has actual values; no need "
                      "to trim")
            break
        log.debug(f"Column {get_column_letter(col)} had no actual "
                  "values; deleting")
        sheet.delete_cols(col)

    # and now from the bottom-most row
    for row in range(sheet.max_row, 1, -1):
        values = [
            sheet.cell(row, col).value is not None
            for col in range(1, sheet.max_column + 1)
        ]
        if sum(values):
            log.debug("Last row of data area has actual values; no need to "
                      "trim")
            break
        log.debug(f"Row {row} had no actual values; deleting")
        sheet.delete_rows(row)

    return sheet


def write_tsv(xlfile, tsvfile, sheetname=None, maxcols=None,
              overwrite=False, delim=None):
    """
    Write a single Excel file to a named tab-delimited output file
    """
    if os.path.exists(tsvfile) and not overwrite:
        raise RuntimeError(f"File '{tsvfile}' exists and overwrite=False")

    if not delim:
        delim = DEFAULT_DELIM

    import csv
    import openpyxl
    from openpyxl.utils.cell import get_column_letter

    wb = openpyxl.load_workbook(xlfile, data_only=True)
    log.debug(f"Opened workbook '{xlfile}'.")
    log.debug(f"Workboook has {len(wb.sheetnames)} worksheet(s): "
              f"{', '.join(wb.sheetnames)}")

    # default to the first sheet if one isn't specified
    sheet = wb[sheetname if sheetname else wb.sheetnames[0]]
    log.info(f"Selecting sheet '{sheet.title}' from workbook '{xlfile}'")
    log.debug(f"Worksheet range: {sheet.calculate_dimension()}")

    if maxcols:
        maxcols = int(maxcols)
        log.info(f"Got maxcols={maxcols}; discarding all columns after "
                 f"{get_column_letter(maxcols)}")
        # only take this many columns and discard the rest
        for col in range(sheet.max_column, maxcols, - 1):
            sheet.delete_cols(col)
        log.debug(f"New worksheet range: {sheet.calculate_dimension()}")

    # trim "data area" by removing non-data columns and rows
    sheet = trim_data_area(sheet)

    class UnquotedTsv(csv.unix_dialect):
        # FIXME: unnecessary, except to prevent the library from erroring out
        escapechar = '\\'
        quoting = csv.QUOTE_NONE

    with open(tsvfile, 'w', newline='') as tsv:
        # without `dialect='unix'`, you get CRLF line endings
        writer = csv.writer(tsv, delimiter=delim, dialect=UnquotedTsv)

        #rownum = 0
        for row in sheet.iter_rows(values_only=True):
            # trim leading/trailing whitespace
            row = [
                x.strip() if isinstance(x, str) else x
                for x in row
            ]
            #rownum += 1
            #log.debug(f"Working on row #{rownum} from '{tsvfile}.'")
            writer.writerow(row)

        log.info(f"Wrote {sheet.max_row} rows of {sheet.max_column} columns "
                 f"to '{tsvfile}'")


@use_config
def write_tsvs(xlfiles=None, outputdir=None, sheetname=None, maxcols=None,
               delim=None, config=None):
    """
    Given a list of Excel files (or tuples w/ dest. filename), write as TSVs

    :param list xlfiles: list containing pathnames to Excel files to read; or,
        a list of tuples, where the second element is the destination filename
    :param str outputdir: the ouput directory
    :param str sheetname: which worksheet (tab) to use from the workbook;
        primarily for testing from the command line
    :param str maxcols: only include this many columns; primarily for testing
    :param dict config: dict of config values if not using ``@use_config``
        decorator function
    """
    import os
    from collections.abc import Iterable

    if not outputdir:
        outputdir = config['site']['deploydatadir']
        log.info(f"No 'outputdir' given, using default of '{outputdir}'")

    if os.path.exists(outputdir):
        if not os.path.isdir(outputdir):
            raise RuntimeError(f"File '{outputdir}' exists (and is not a "
                                "directory)")
    else:
        log.info(f"Output directory '{outputdir}' doesn't exist; creating")
        os.makedirs(outputdir, exist_ok=True)

    if xlfiles:
        # if it's a single string, and the pathname exists
        if isinstance(xlfiles, str) and os.path.exists(xlfiles):
            basename = os.path.splitext(os.path.split(xlfiles)[1])[0]
            tsvfiles = [(
                xlfiles,   # which is actually just a single string
                os.path.join(outputdir, basename + ".tsv"),
                sheetname
            )]
        # it's a list of Excel files
        elif isinstance(xlfiles, Iterable):
            if len(xlfiles) == 0:
                raise RuntimeError("Empty list 'xlfiles'")
            if not isinstance(xlfiles[0], Iterable):
                xlfiles = [
                    (x, os.path.basename(x), sheetname)
                    for x in xlfiles
                ]
        else:
            raise RuntimeError("Bad value for 'xlfiles'")
    else:
        # read Excel filenames from data.toml
        basename = config['data']['artifacts']['basename']
        tsvfiles = []
        for d in config['data']['datasources']:
            infile = os.path.join(
                config['site']['deploydatadir'],
                config['data']['datasources'][d]['outfilename']
            )
            outfile = os.path.splitext(infile)[0] + '.tsv'

            tsvfiles.append((
                infile, outfile,
                sheetname if sheetname else 'Sheet1'
            ))

    log.debug(f"Files to process: {tsvfiles}")
    for t in tsvfiles:
        # first 3 args (xlsfile, tsvfile, sheetname), come from the tuple
        write_tsv(*t, maxcols=maxcols, overwrite=True, delim=delim)


if __name__ == '__main__':
    import sys

    if '-h' in sys.argv or '--help' in sys.argv:
        print("""
  usage:
    python -m lib.data [xlfile] [outputdir] [sheetname] [maxcols] [delim]
""")
        sys.exit()

    write_tsvs(*sys.argv[1:])
